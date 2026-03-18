from flask import Blueprint, render_template, send_file, request, flash, redirect, url_for
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import db, Rendicion, User, ItemRendicion
from sqlalchemy import func, extract
from utils.decorators import permission_required

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')


@reportes_bp.route('/')
@login_required
@permission_required('aprobador')
def index():
    """Página principal de reportes"""
    return render_template('reportes/index.html')


@reportes_bp.route('/rendiciones')
@login_required
@permission_required('aprobador')
def rendiciones():
    """Reporte de rendiciones con filtros"""
    # Obtener parámetros
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    estado = request.args.get('estado')
    usuario_id = request.args.get('usuario_id', type=int)
    
    # Query base
    query = Rendicion.query
    
    # Aplicar filtros
    if fecha_desde:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
        query = query.filter(Rendicion.fecha_creacion >= fecha_desde_obj)
    
    if fecha_hasta:
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        query = query.filter(Rendicion.fecha_creacion <= fecha_hasta_obj)
    
    if estado:
        query = query.filter_by(estado=estado)
    
    if usuario_id:
        query = query.filter_by(usuario_id=usuario_id)
    
    # Obtener resultados
    rendiciones = query.order_by(Rendicion.fecha_creacion.desc()).all()
    
    # Calcular totales
    total_rendiciones = len(rendiciones)
    monto_total = sum(r.monto_total for r in rendiciones)
    
    # Obtener lista de usuarios para el filtro
    usuarios = User.query.filter_by(activo=True).order_by(User.nombre).all()
    
    return render_template('reportes/rendiciones.html',
                         rendiciones=rendiciones,
                         usuarios=usuarios,
                         total_rendiciones=total_rendiciones,
                         monto_total=monto_total,
                         fecha_desde=fecha_desde,
                         fecha_hasta=fecha_hasta,
                         estado=estado,
                         usuario_id=usuario_id)


@reportes_bp.route('/rendiciones/exportar')
@login_required
@permission_required('aprobador')
def exportar_rendiciones():
    """Exportar rendiciones a Excel"""
    # Obtener parámetros (mismos filtros que la vista)
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    estado = request.args.get('estado')
    usuario_id = request.args.get('usuario_id', type=int)
    
    # Query base
    query = Rendicion.query
    
    # Aplicar filtros
    if fecha_desde:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
        query = query.filter(Rendicion.fecha_creacion >= fecha_desde_obj)
    
    if fecha_hasta:
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
        query = query.filter(Rendicion.fecha_creacion <= fecha_hasta_obj)
    
    if estado:
        query = query.filter_by(estado=estado)
    
    if usuario_id:
        query = query.filter_by(usuario_id=usuario_id)
    
    rendiciones = query.order_by(Rendicion.fecha_creacion.desc()).all()
    
    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendiciones"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'Número', 'Usuario', 'Fecha Creación', 'Fecha Inicio', 'Fecha Fin',
        'Descripción', 'Estado', 'Monto Total', 'Aprobador', 'Fecha Aprobación'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, rendicion in enumerate(rendiciones, 2):
        ws.cell(row=row, column=1, value=rendicion.numero_rendicion).border = border
        ws.cell(row=row, column=2, value=rendicion.usuario.nombre).border = border
        ws.cell(row=row, column=3, value=rendicion.fecha_creacion.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row, column=4, value=rendicion.fecha_inicio.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=5, value=rendicion.fecha_fin.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=6, value=rendicion.descripcion or '').border = border
        ws.cell(row=row, column=7, value=rendicion.get_estado_display()).border = border
        ws.cell(row=row, column=8, value=float(rendicion.monto_total)).border = border
        ws.cell(row=row, column=8).number_format = '$#,##0'
        ws.cell(row=row, column=9, value=rendicion.aprobador.nombre if rendicion.aprobador else '').border = border
        ws.cell(row=row, column=10, value=rendicion.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if rendicion.fecha_aprobacion else '').border = border
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generar nombre de archivo
    filename = f"rendiciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reportes_bp.route('/rendiciones/<int:id>/exportar-detalle')
@login_required
def exportar_detalle_rendicion(id):
    """Exportar detalle de una rendición a Excel"""
    rendicion = Rendicion.query.get_or_404(id)
    
    # Verificar permisos
    if not current_user.can_approve() and rendicion.usuario_id != current_user.id:
        flash('No tienes permiso para exportar esta rendición', 'error')
        return redirect(url_for('rendiciones.index'))
    
    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle Rendición"
    
    # Estilos
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Información de la rendición
    ws['A1'] = 'DETALLE DE RENDICIÓN'
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    row = 3
    info = [
        ('Número:', rendicion.numero_rendicion),
        ('Usuario:', rendicion.usuario.nombre),
        ('Estado:', rendicion.get_estado_display()),
        ('Fecha Creación:', rendicion.fecha_creacion.strftime('%d/%m/%Y %H:%M')),
        ('Periodo:', f"{rendicion.fecha_inicio.strftime('%d/%m/%Y')} - {rendicion.fecha_fin.strftime('%d/%m/%Y')}"),
        ('Descripción:', rendicion.descripcion or ''),
        ('Monto Total:', f"${rendicion.monto_total:,.0f}")
    ]
    
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    # Saltar una fila
    row += 2
    
    # Encabezados de items
    headers = ['Fecha', 'Tipo Gasto', 'Descripción', 'Proveedor', 'N° Doc', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row += 1
    
    # Items
    items = rendicion.items.all()
    for item in items:
        ws.cell(row=row, column=1, value=item.fecha_gasto.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=2, value=item.tipo_gasto).border = border
        ws.cell(row=row, column=3, value=item.descripcion).border = border
        ws.cell(row=row, column=4, value=item.proveedor or '').border = border
        ws.cell(row=row, column=5, value=item.numero_documento or '').border = border
        ws.cell(row=row, column=6, value=float(item.monto)).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0'
        row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rendicion_{rendicion.numero_rendicion}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reportes_bp.route('/estadisticas')
@login_required
@permission_required('aprobador')
def estadisticas():
    """Estadísticas generales del sistema"""
    # Rango de fechas
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    
    # Por defecto, últimos 3 meses
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')
    
    fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
    fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
    
    # Rendiciones por estado
    rendiciones_por_estado = db.session.query(
        Rendicion.estado,
        func.count(Rendicion.id).label('count'),
        func.sum(Rendicion.monto_total).label('monto')
    ).filter(
        Rendicion.fecha_creacion.between(fecha_desde_obj, fecha_hasta_obj)
    ).group_by(Rendicion.estado).all()
    
    # Rendiciones por mes
    rendiciones_por_mes = db.session.query(
        func.date_format(Rendicion.fecha_creacion, '%Y-%m').label('mes'),
        func.count(Rendicion.id).label('count'),
        func.sum(Rendicion.monto_total).label('monto')
    ).filter(
        Rendicion.fecha_creacion.between(fecha_desde_obj, fecha_hasta_obj)
    ).group_by('mes').order_by('mes').all()
    
    # Top usuarios por monto
    top_usuarios = db.session.query(
        User.nombre,
        func.count(Rendicion.id).label('count'),
        func.sum(Rendicion.monto_total).label('monto')
    ).join(Rendicion).filter(
        Rendicion.fecha_creacion.between(fecha_desde_obj, fecha_hasta_obj),
        Rendicion.estado.in_(['aprobada', 'pagada'])
    ).group_by(User.id).order_by(func.sum(Rendicion.monto_total).desc()).limit(10).all()
    
    # Gastos por tipo
    gastos_por_tipo = db.session.query(
        ItemRendicion.tipo_gasto,
        func.count(ItemRendicion.id).label('count'),
        func.sum(ItemRendicion.monto).label('monto')
    ).join(Rendicion).filter(
        Rendicion.fecha_creacion.between(fecha_desde_obj, fecha_hasta_obj)
    ).group_by(ItemRendicion.tipo_gasto).order_by(func.sum(ItemRendicion.monto).desc()).all()
    
    return render_template('reportes/estadisticas.html',
                         fecha_desde=fecha_desde,
                         fecha_hasta=fecha_hasta,
                         rendiciones_por_estado=rendiciones_por_estado,
                         rendiciones_por_mes=rendiciones_por_mes,
                         top_usuarios=top_usuarios,
                         gastos_por_tipo=gastos_por_tipo)