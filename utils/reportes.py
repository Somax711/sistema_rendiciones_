"""
Utilidades para generación de reportes
Sistema de Rendiciones - Primar
"""

import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import db, Rendicion, User, ItemRendicion
from sqlalchemy import and_, or_, func


def generar_reporte_rendiciones(formato='xlsx', fecha_desde=None, fecha_hasta=None, estado=None, usuario_id=None):
    """
    Genera reporte de rendiciones con filtros
    
    Args:
        formato: Formato del reporte (xlsx, csv, pdf)
        fecha_desde: Fecha de inicio del filtro
        fecha_hasta: Fecha de fin del filtro
        estado: Estado de las rendiciones
        usuario_id: ID del usuario
    
    Returns:
        Tupla (ruta_archivo, nombre_archivo)
    """
    # Construir query con filtros
    query = Rendicion.query
    
    if fecha_desde:
        query = query.filter(Rendicion.fecha_rendicion >= fecha_desde)
    
    if fecha_hasta:
        query = query.filter(Rendicion.fecha_rendicion <= fecha_hasta)
    
    if estado:
        query = query.filter(Rendicion.estado == estado)
    
    if usuario_id:
        query = query.filter(Rendicion.usuario_id == usuario_id)
    
    # Ordenar por fecha descendente
    rendiciones = query.order_by(Rendicion.fecha_rendicion.desc()).all()
    
    if formato == 'xlsx':
        return _generar_excel_rendiciones(rendiciones)
    elif formato == 'csv':
        return _generar_csv_rendiciones(rendiciones)
    elif formato == 'pdf':
        return _generar_pdf_rendiciones(rendiciones)
    else:
        raise ValueError(f"Formato no soportado: {formato}")


def _generar_excel_rendiciones(rendiciones):
    """
    Genera archivo Excel con las rendiciones
    
    Args:
        rendiciones: Lista de objetos Rendicion
    
    Returns:
        Tupla (ruta_archivo, nombre_archivo)
    """
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rendiciones"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "N° Rendición",
        "Usuario",
        "Fecha",
        "Descripción",
        "Monto Total",
        "Estado",
        "Fecha Creación",
        "Fecha Aprobación"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos
    for row, rendicion in enumerate(rendiciones, 2):
        ws.cell(row=row, column=1).value = rendicion.numero_rendicion
        ws.cell(row=row, column=2).value = rendicion.usuario.nombre
        ws.cell(row=row, column=3).value = rendicion.fecha_rendicion.strftime('%d/%m/%Y')
        ws.cell(row=row, column=4).value = rendicion.descripcion
        ws.cell(row=row, column=5).value = float(rendicion.monto_total)
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        ws.cell(row=row, column=6).value = rendicion.estado.upper()
        ws.cell(row=row, column=7).value = rendicion.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        
        if rendicion.fecha_aprobacion:
            ws.cell(row=row, column=8).value = rendicion.fecha_aprobacion.strftime('%d/%m/%Y %H:%M')
        
        # Aplicar bordes
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar anchos de columna
    column_widths = [15, 25, 12, 40, 15, 15, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Agregar hoja de resumen
    ws_resumen = wb.create_sheet("Resumen")
    
    # Resumen por estado
    ws_resumen.cell(row=1, column=1).value = "RESUMEN POR ESTADO"
    ws_resumen.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    estados = {}
    for rendicion in rendiciones:
        estado = rendicion.estado
        if estado not in estados:
            estados[estado] = {'cantidad': 0, 'monto': 0}
        estados[estado]['cantidad'] += 1
        estados[estado]['monto'] += float(rendicion.monto_total)
    
    row = 3
    ws_resumen.cell(row=row, column=1).value = "Estado"
    ws_resumen.cell(row=row, column=2).value = "Cantidad"
    ws_resumen.cell(row=row, column=3).value = "Monto Total"
    
    for col in range(1, 4):
        cell = ws_resumen.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row += 1
    for estado, datos in estados.items():
        ws_resumen.cell(row=row, column=1).value = estado.upper()
        ws_resumen.cell(row=row, column=2).value = datos['cantidad']
        ws_resumen.cell(row=row, column=3).value = datos['monto']
        ws_resumen.cell(row=row, column=3).number_format = '$#,##0.00'
        row += 1
    
    # Guardar archivo temporal
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Rendiciones_{fecha}.xlsx"
    
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    
    wb.save(archivo_path)
    
    return archivo_path, nombre_archivo


def _generar_csv_rendiciones(rendiciones):
    """
    Genera archivo CSV con las rendiciones
    
    Args:
        rendiciones: Lista de objetos Rendicion
    
    Returns:
        Tupla (ruta_archivo, nombre_archivo)
    """
    import csv
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Rendiciones_{fecha}.csv"
    
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    
    with open(archivo_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        
        # Headers
        writer.writerow([
            'N° Rendición',
            'Usuario',
            'Fecha',
            'Descripción',
            'Monto Total',
            'Estado',
            'Fecha Creación',
            'Fecha Aprobación'
        ])
        
        # Datos
        for rendicion in rendiciones:
            writer.writerow([
                rendicion.numero_rendicion,
                rendicion.usuario.nombre,
                rendicion.fecha_rendicion.strftime('%d/%m/%Y'),
                rendicion.descripcion,
                float(rendicion.monto_total),
                rendicion.estado.upper(),
                rendicion.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                rendicion.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if rendicion.fecha_aprobacion else ''
            ])
    
    return archivo_path, nombre_archivo


def _generar_pdf_rendiciones(rendiciones):
    """
    Genera archivo PDF con las rendiciones
    
    Args:
        rendiciones: Lista de objetos Rendicion
    
    Returns:
        Tupla (ruta_archivo, nombre_archivo)
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Rendiciones_{fecha}.pdf"
    
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    
    # Crear documento
    doc = SimpleDocTemplate(archivo_path, pagesize=landscape(letter))
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Título
    title = Paragraph("REPORTE DE RENDICIONES", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Información del reporte
    fecha_reporte = datetime.now().strftime('%d/%m/%Y %H:%M')
    info = Paragraph(f"Fecha de generación: {fecha_reporte}", styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Tabla de datos
    data = [[
        'N° Rendición',
        'Usuario',
        'Fecha',
        'Descripción',
        'Monto Total',
        'Estado'
    ]]
    
    for rendicion in rendiciones:
        data.append([
            rendicion.numero_rendicion,
            rendicion.usuario.nombre[:20],  # Limitar longitud
            rendicion.fecha_rendicion.strftime('%d/%m/%Y'),
            rendicion.descripcion[:30],  # Limitar longitud
            f"${float(rendicion.monto_total):,.2f}",
            rendicion.estado.upper()
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[1.2*inch, 1.8*inch, 1*inch, 2.5*inch, 1.2*inch, 1*inch])
    
    # Estilo de tabla
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    return archivo_path, nombre_archivo


def generar_reporte_usuarios(formato='xlsx'):
    """
    Genera reporte de usuarios del sistema
    
    Args:
        formato: Formato del reporte (xlsx, csv)
    
    Returns:
        Tupla (ruta_archivo, nombre_archivo)
    """
    usuarios = User.query.order_by(User.nombre).all()
    
    if formato == 'xlsx':
        return _generar_excel_usuarios(usuarios)
    elif formato == 'csv':
        return _generar_csv_usuarios(usuarios)
    else:
        raise ValueError(f"Formato no soportado: {formato}")


def _generar_excel_usuarios(usuarios):
    """Genera reporte Excel de usuarios"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Headers
    headers = ["ID", "Nombre", "Email", "Rol", "Estado", "Fecha Registro", "MFA Habilitado"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    # Datos
    for row, usuario in enumerate(usuarios, 2):
        ws.cell(row=row, column=1).value = usuario.id
        ws.cell(row=row, column=2).value = usuario.nombre
        ws.cell(row=row, column=3).value = usuario.email
        ws.cell(row=row, column=4).value = usuario.rol.upper()
        ws.cell(row=row, column=5).value = "ACTIVO" if usuario.activo else "INACTIVO"
        ws.cell(row=row, column=6).value = usuario.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row, column=7).value = "SÍ" if usuario.mfa_habilitado else "NO"
    
    # Guardar
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Usuarios_{fecha}.xlsx"
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    wb.save(archivo_path)
    
    return archivo_path, nombre_archivo


def _generar_csv_usuarios(usuarios):
    """Genera reporte CSV de usuarios"""
    import csv
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Usuarios_{fecha}.csv"
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    
    with open(archivo_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['ID', 'Nombre', 'Email', 'Rol', 'Estado', 'Fecha Registro', 'MFA'])
        
        for usuario in usuarios:
            writer.writerow([
                usuario.id,
                usuario.nombre,
                usuario.email,
                usuario.rol.upper(),
                "ACTIVO" if usuario.activo else "INACTIVO",
                usuario.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                "SÍ" if usuario.mfa_habilitado else "NO"
            ])
    
    return archivo_path, nombre_archivo


def generar_reporte_aprobaciones(formato='xlsx', fecha_desde=None, fecha_hasta=None):
    """
    Genera reporte de aprobaciones
    
    Args:
        formato: Formato del reporte
        fecha_desde: Fecha inicio
        fecha_hasta: Fecha fin
    
    Returns:
        Tupla (ruta_archivo, nombre_archivo)
    """
    # Query de rendiciones aprobadas o rechazadas
    query = Rendicion.query.filter(
        Rendicion.estado.in_(['aprobada', 'rechazada', 'pagada'])
    )
    
    if fecha_desde:
        query = query.filter(Rendicion.fecha_aprobacion >= fecha_desde)
    
    if fecha_hasta:
        query = query.filter(Rendicion.fecha_aprobacion <= fecha_hasta)
    
    rendiciones = query.order_by(Rendicion.fecha_aprobacion.desc()).all()
    
    if formato == 'xlsx':
        return _generar_excel_aprobaciones(rendiciones)
    elif formato == 'csv':
        return _generar_csv_aprobaciones(rendiciones)
    else:
        raise ValueError(f"Formato no soportado: {formato}")


def _generar_excel_aprobaciones(rendiciones):
    """Genera reporte Excel de aprobaciones"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Aprobaciones"
    
    headers = [
        "N° Rendición", "Usuario", "Fecha Rendición", "Monto",
        "Estado", "Aprobador", "Fecha Aprobación", "Comentarios"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    for row, rendicion in enumerate(rendiciones, 2):
        ws.cell(row=row, column=1).value = rendicion.numero_rendicion
        ws.cell(row=row, column=2).value = rendicion.usuario.nombre
        ws.cell(row=row, column=3).value = rendicion.fecha_rendicion.strftime('%d/%m/%Y')
        ws.cell(row=row, column=4).value = float(rendicion.monto_total)
        ws.cell(row=row, column=5).value = rendicion.estado.upper()
        
        if rendicion.aprobador_id:
            ws.cell(row=row, column=6).value = rendicion.aprobador.nombre
        
        if rendicion.fecha_aprobacion:
            ws.cell(row=row, column=7).value = rendicion.fecha_aprobacion.strftime('%d/%m/%Y %H:%M')
        
        ws.cell(row=row, column=8).value = rendicion.comentarios_aprobacion or ""
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Aprobaciones_{fecha}.xlsx"
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    wb.save(archivo_path)
    
    return archivo_path, nombre_archivo


def _generar_csv_aprobaciones(rendiciones):
    """Genera reporte CSV de aprobaciones"""
    import csv
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Aprobaciones_{fecha}.csv"
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    
    with open(archivo_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([
            'N° Rendición', 'Usuario', 'Fecha', 'Monto',
            'Estado', 'Aprobador', 'Fecha Aprobación', 'Comentarios'
        ])
        
        for rendicion in rendiciones:
            writer.writerow([
                rendicion.numero_rendicion,
                rendicion.usuario.nombre,
                rendicion.fecha_rendicion.strftime('%d/%m/%Y'),
                float(rendicion.monto_total),
                rendicion.estado.upper(),
                rendicion.aprobador.nombre if rendicion.aprobador_id else '',
                rendicion.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if rendicion.fecha_aprobacion else '',
                rendicion.comentarios_aprobacion or ''
            ])
    
    return archivo_path, nombre_archivo


def generar_reporte_gastos(formato='xlsx', fecha_desde=None, fecha_hasta=None, usuario_id=None):
    """
    Genera reporte detallado de gastos por categoría
    
    Args:
        formato: Formato del reporte
        fecha_desde: Fecha inicio
        fecha_hasta: Fecha fin
        usuario_id: ID del usuario
    
    Returns:
        Tupla (ruta_archivo, nombre_archivo)
    """
    query = db.session.query(
        ItemRendicion,
        Rendicion,
        User
    ).join(
        Rendicion, ItemRendicion.rendicion_id == Rendicion.id
    ).join(
        User, Rendicion.usuario_id == User.id
    )
    
    if fecha_desde:
        query = query.filter(Rendicion.fecha_rendicion >= fecha_desde)
    
    if fecha_hasta:
        query = query.filter(Rendicion.fecha_rendicion <= fecha_hasta)
    
    if usuario_id:
        query = query.filter(Rendicion.usuario_id == usuario_id)
    
    items = query.order_by(Rendicion.fecha_rendicion.desc()).all()
    
    if formato == 'xlsx':
        return _generar_excel_gastos(items)
    elif formato == 'csv':
        return _generar_csv_gastos(items)
    else:
        raise ValueError(f"Formato no soportado: {formato}")


def _generar_excel_gastos(items):
    """Genera reporte Excel detallado de gastos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos Detallados"
    
    headers = [
        "Fecha", "N° Rendición", "Usuario", "Categoría",
        "Descripción", "Cantidad", "Valor Unitario", "Monto Total", "Estado"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    for row, (item, rendicion, usuario) in enumerate(items, 2):
        ws.cell(row=row, column=1).value = rendicion.fecha_rendicion.strftime('%d/%m/%Y')
        ws.cell(row=row, column=2).value = rendicion.numero_rendicion
        ws.cell(row=row, column=3).value = usuario.nombre
        ws.cell(row=row, column=4).value = item.categoria or "Sin categoría"
        ws.cell(row=row, column=5).value = item.descripcion
        ws.cell(row=row, column=6).value = item.cantidad
        ws.cell(row=row, column=7).value = float(item.valor_unitario)
        ws.cell(row=row, column=8).value = float(item.monto_total)
        ws.cell(row=row, column=9).value = rendicion.estado.upper()
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Gastos_{fecha}.xlsx"
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    wb.save(archivo_path)
    
    return archivo_path, nombre_archivo


def _generar_csv_gastos(items):
    """Genera reporte CSV detallado de gastos"""
    import csv
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Gastos_{fecha}.csv"
    temp_dir = tempfile.gettempdir()
    archivo_path = os.path.join(temp_dir, nombre_archivo)
    
    with open(archivo_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([
            'Fecha', 'N° Rendición', 'Usuario', 'Categoría',
            'Descripción', 'Cantidad', 'Valor Unitario', 'Monto Total', 'Estado'
        ])
        
        for item, rendicion, usuario in items:
            writer.writerow([
                rendicion.fecha_rendicion.strftime('%d/%m/%Y'),
                rendicion.numero_rendicion,
                usuario.nombre,
                item.categoria or "Sin categoría",
                item.descripcion,
                item.cantidad,
                float(item.valor_unitario),
                float(item.monto_total),
                rendicion.estado.upper()
            ])
    
    return archivo_path, nombre_archivo